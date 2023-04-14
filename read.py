import json
from datetime import datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from title import Title

book = openpyxl.open("5MN.xlsx")

title = book.worksheets[0]
section1 = book.worksheets[1]
section2 = book.worksheets[2]
section3 = book.worksheets[3]
codeColumn = None
mnemoCodeColumn = None
nameColumn = None
regionRow = None
districtRow = None
cadCostRow = None

def getReportDate(cell: Worksheet.cell):
    try:
        value = str(cell.value)
        if value.__contains__("О НАЛОГОВОЙ БАЗЕ И СТРУКТУРЕ НАЧИСЛЕНИЙ ПО МЕСТНЫМ НАЛОГАМ"):
            year = '1/1/' + cell.value[-8:][:4]
            try:
                return datetime.strptime(year, '%m/%d/%Y')
            except ValueError:
                print("Not a date" + cell.value)
    except ValueError:
        print("Not a string")

def findCellRow(cell: Worksheet.cell, containedValue: str):
    try:
        value = str(cell.value)
        if value == containedValue:
            return cell.row
    except ValueError:
        print("Not a string")
def findCellColumn(cell: Worksheet.cell, containedValue: str):
    try:
        value = str(cell.value)
        if value == containedValue:
            return cell.column
    except ValueError:
        print("Not a string")


titleModel = Title()
for row in range(1, title.max_row + 1):
    for col in range(1, title.max_column + 1):
        cell = title.cell(row=row, column=col)
        if cell.value is not None:
            year = getReportDate(cell)
            if year is not None:
                titleModel.year = year

            if codeColumn is None:
                codeColumn = findCellColumn(cell, 'Код')
            if nameColumn is None:
                nameColumn = findCellColumn(cell, 'Наименование')
            if regionRow is None:
                regionRow = findCellRow(cell, 'Республика, край, область, автономное образование, город')
            if districtRow is None:
                districtRow = findCellRow(cell, 'Муниципальное образование')

            if codeColumn is not None and regionRow is not None:
                titleModel.regionCode = title.cell(row=regionRow, column=codeColumn).value
                titleModel.regionName = title.cell(row=regionRow, column=nameColumn).value

            if codeColumn is not None and districtRow is not None:
                titleModel.districtCode = title.cell(row=districtRow, column=codeColumn).value
                titleModel.districtName = title.cell(row=districtRow, column=nameColumn).value



jsonStr = json.dumps(titleModel.__dict__, indent=4, sort_keys=True, default=str)
# print(jsonStr)

for row in range(1, section2.max_row):
    for col in range(1, section2.max_column):
        cell = section2.cell(row=row, column=col)
        if cell.value is not None:
            if mnemoCodeColumn is None:
                mnemoCodeColumn = findCellColumn(cell, 'Значение показателя')

            if cadCostRow is None:
                cadCostRow = findCellRow(cell, '4.  Кадастровая стоимость')

            if mnemoCodeColumn is not None and cadCostRow is not None:
                print(section2.cell(row=cadCostRow, column=mnemoCodeColumn).value)